from fastapi import APIRouter, Depends, Query, Response
from fastapi.responses import StreamingResponse
from typing import Optional
from routes.auth import require_auth
from cache import cache
from datetime import datetime
import pandas as pd
import io
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])


def _pct(num, den):
    """Calcula el porcentaje de forma segura."""
    if not den or den == 0:
        return 0.0
    return round(float(num) / float(den) * 100, 1)


def apply_excel_style(worksheet, sheet_name="Sheet1"):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    # Styling constants
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Blue-900
    header_font = Font(color='FFFFFF', bold=True, size=11)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    # Format Headers (First Row)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
        
    # Auto-adjust column width and alternate row shading
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
                
                # Apply borders and padding-like styling
                if cell.row > 1:
                    cell.border = border
                    if cell.row % 2 == 0:
                        cell.fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            except:
                pass
        adjusted_width = (max_length + 4)
        worksheet.column_dimensions[column].width = min(adjusted_width, 50) # Cap width


@router.get("/export")
async def export_leads(
    search: Optional[str] = Query(None),
    base: Optional[str] = Query(None),
    programa: Optional[str] = Query(None),
    nivel: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    no_util: Optional[bool] = Query(False),
    _user: str = Depends(require_auth),
):
    from database import fetch_all
    
    where_clauses = ["1=1"]
    args = []

    if no_util:
        where_clauses.append("(descrip_cat ILIKE '%no util%' OR descrip_cat ILIKE '%descarte%')")
    
    if search:
        where_clauses.append(f"(txtnombreapellido ILIKE ${len(args)+1} OR emlmail ILIKE ${len(args)+1} OR teltelefono ILIKE ${len(args)+1})")
        args.append(f"%{search}%")
        
    if base:
        where_clauses.append(f"base = ${len(args)+1}")
        args.append(base)
        
    if programa:
        where_clauses.append(f"txtprogramainteres ILIKE ${len(args)+1}")
        args.append(f"%{programa}%")
        
    if estado:
        where_clauses.append(f"ultima_mejor_subcat_string = ${len(args)+1}")
        args.append(estado)
        
    if fecha_inicio:
        where_clauses.append(f"fecha_a_utilizar >= ${len(args)+1}")
        args.append(fecha_inicio)
        
    if fecha_fin:
        where_clauses.append(f"fecha_a_utilizar <= ${len(args)+1}")
        args.append(fecha_fin)
        
    if nivel and nivel.upper() != "TODOS":
        target_nivel = nivel.upper()
        data_cache = await cache.get_all()
        programs_of_level = [p["programa"] for p in data_cache.get("merged_programs", []) if p.get("nivel") == target_nivel]
        
        if programs_of_level:
            placeholders = ",".join(f"${len(args)+i+1}" for i in range(len(programs_of_level)))
            where_clauses.append(f"UPPER(TRIM(txtprogramainteres)) IN ({placeholders})")
            args.extend(programs_of_level)
        else:
            where_clauses.append("1=0")
        
    where_sql = " AND ".join(where_clauses)
    
    data_query = f"""
        SELECT 
            idinterno AS "ID INTERNO", 
            txtnombreapellido AS "NOMBRE Y APELLIDO", 
            emlmail AS "EMAIL", 
            teltelefono AS "TELEFONO", 
            txtprogramainteres AS "PROGRAMA INTERES", 
            base AS "BASE DE DATOS",
            ultima_mejor_subcat_string AS "ESTADO GESTION",
            descrip_subcat AS "SUBCATEGORIA",
            fecha_a_utilizar AS "FECHA ACTIVIDAD"
        FROM dim_contactos
        WHERE {where_sql}
        ORDER BY fecha_a_utilizar DESC NULLS LAST
    """
    
    rows = await fetch_all(data_query, *args)
    print(f"[Export] Exporting {len(rows)} leads for user. Filters: no_util={no_util}, nivel={nivel}, search={search}")
    df = pd.DataFrame(rows)
    
    # Generate Excel in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Leads')
        apply_excel_style(writer.sheets['Leads'])

    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="Expert_Leads_Report.xlsx"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.get("/export-admisiones")
async def export_admisiones(
    nivel: Optional[str] = Query(None),
    _user: str = Depends(require_auth)
):
    data = await cache.get_all()
    merged = data.get("merged_programs", [])
    
    if nivel and nivel.upper() != "TODOS":
        nivel = nivel.upper()
        merged = [p for p in merged if p.get("nivel") == nivel]
    
    # Clean data for export
    export_data = []
    for p in merged:
        export_data.append({
            "PROGRAMA": p.get("programa"),
            "NIVEL": p.get("nivel"),
            "AREA": p.get("area"),
            "LEADS": p.get("leads"),
            "SOLICITADOS": p.get("solicitados"),
            "ADMITIDOS": p.get("admitidos"),
            "PAGADOS": p.get("pagados"),
            "META": p.get("meta"),
            "CUMPLIMIENTO %": _pct(p.get("pagados", 0), p.get("meta", 0))
        })
    
    df = pd.DataFrame(export_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Admisiones')
        apply_excel_style(writer.sheets['Admisiones'])

    output.seek(0)
    filename = f"Reporte_Admisiones_{nivel if nivel else 'GLOBAL'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.get("/kpis")
async def get_kpis(nivel: Optional[str] = Query(None), _user: str = Depends(require_auth)):
    data = await cache.get_all()
    if not nivel or nivel.upper() == "TODOS":
        totals = data.get("totals", {})
        return {
            "total_leads": data.get("total_leads", 0),
            "en_gestion": data.get("en_gestion", 0),
            "op_venta": data.get("op_venta", 0),
            "solicitados": totals.get("solicitados", 0),
            "admitidos": totals.get("admitidos", 0),
            "pagados": totals.get("pagados", 0),
            "metas": totals.get("metas", 0),
            "matriculados": totals.get("pagados", 0),
            "proceso_pago": data.get("proceso_pago", 0),
            "no_util_total": data.get("no_util_total", 0),
            "fecha_actualizacion": data.get("fecha_actualizacion", ""),
            "trends": data.get("trends", {})
        }
    
    # Filter and re-aggregate
    nivel = nivel.upper()
    programs = [p for p in data.get("merged_programs", []) if str(p.get("nivel", "")).upper() == nivel]
    
    total_leads = sum(p.get("leads", 0) for p in programs)
    en_gestion = sum(p.get("en_gestion", 0) for p in programs)
    op_venta = sum(p.get("op_venta", 0) for p in programs)
    solicitados = sum(p.get("solicitados", 0) for p in programs)
    admitidos = sum(p.get("admitidos", 0) for p in programs)
    pagados = sum(p.get("pagados", 0) for p in programs)
    metas = sum(p.get("meta", 0) for p in programs)
    proceso_pago = sum(p.get("proceso_pago", 0) for p in programs)
    no_util_total = sum(p.get("no_util", 0) for p in programs)

    return {
        "total_leads": total_leads,
        "en_gestion": en_gestion,
        "op_venta": op_venta,
        "solicitados": solicitados,
        "admitidos": admitidos,
        "pagados": pagados,
        "metas": metas,
        "matriculados": pagados,
        "proceso_pago": proceso_pago,
        "no_util_total": no_util_total,
        "fecha_actualizacion": data.get("fecha_actualizacion", ""),
        "trends": {} # Trends are complex to re-calc on the fly, keeping empty for filtered view
    }


@router.get("/funnel")
async def get_funnel(nivel: Optional[str] = Query(None), _user: str = Depends(require_auth)):
    data = await cache.get_all()
    
    if not nivel or nivel.upper() == "TODOS":
        funnel_data = data.get("funnel", [])
        total_leads = data.get("total_leads", 0)
    else:
        nivel = nivel.upper()
        programs = [p for p in data.get("merged_programs", []) if str(p.get("nivel", "")).upper() == nivel]
        total_leads = sum(p.get("leads", 0) for p in programs)
        en_gestion = sum(p.get("en_gestion", 0) for p in programs)
        op_venta = sum(p.get("op_venta", 0) for p in programs)
        proceso_pago = sum(p.get("proceso_pago", 0) for p in programs)
        pagados = sum(p.get("pagados", 0) for p in programs)
        
        funnel_data = [
            {"stage": "Total Leads", "value": total_leads, "color": "#f59e0b"},
            {"stage": "En Gestión", "value": en_gestion, "color": "#d97706"},
            {"stage": "Oportunidad de Venta", "value": op_venta, "color": "#ea580c"},
            {"stage": "Proceso Pago", "value": proceso_pago, "color": "#dc2626"},
            {"stage": "Matriculados", "value": pagados, "color": "#16a34a"},
        ]
    
    # Calculate percentages for safety
    for f in funnel_data:
        f["percent"] = round(f["value"] / total_leads * 100, 2) if total_leads else 0
        
    return funnel_data


@router.get("/admisiones")
async def get_admisiones(nivel: Optional[str] = Query(None), _user: str = Depends(require_auth)):
    data = await cache.get_all()
    merged = data.get("merged_programs", [])
    
    if nivel and nivel.upper() != "TODOS":
        nivel = nivel.upper()
        merged = [p for p in merged if p.get("nivel") == nivel]
        
        totals = {
            "solicitados": sum(p.get("solicitados", 0) for p in merged),
            "admitidos": sum(p.get("admitidos", 0) for p in merged),
            "pagados": sum(p.get("pagados", 0) for p in merged),
            "metas": sum(p.get("meta", 0) for p in merged),
        }
        return {"programas": merged, "totals": totals, "trends": {}}

    return {
        "programas": merged,
        "totals": data.get("totals", {}),
        "trends": data.get("trends", {}),
    }


@router.get("/export-estados")
async def export_estados(
    nivel: Optional[str] = Query(None),
    _user: str = Depends(require_auth)
):
    data = await cache.get_all()
    merged = data.get("merged_programs", [])
    
    if nivel and nivel.upper() != "TODOS":
        nivel = nivel.upper()
        merged = [p for p in merged if p.get("nivel") == nivel]
    
    # Clean data for export
    export_data = []
    for p in merged:
        export_data.append({
            "PROGRAMA": p.get("programa"),
            "NIVEL": p.get("nivel"),
            "LEADS": p.get("leads"),
            "EN GESTION": p.get("en_gestion"),
            "NO UTIL": p.get("no_util"),
            "OP. VENTA": p.get("op_venta"),
            "PROC. PAGO": p.get("proceso_pago"),
            "SOLICITADOS": p.get("solicitados"),
            "ADMITIDOS": p.get("admitidos"),
            "PAGADOS": p.get("pagados"),
            "META": p.get("meta"),
            "AVANCE %": _pct(p.get("pagados", 0), p.get("meta", 0)),
            "CONVERSION %": _pct(p.get("pagados", 0), p.get("leads", 0))
        })
    
    df = pd.DataFrame(export_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Estados de Gestion')
        apply_excel_style(writer.sheets['Estados de Gestion'])

    output.seek(0)
    filename = f"Reporte_Estados_{nivel if nivel else 'GLOBAL'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )

@router.get("/estados")
async def get_estados(nivel: Optional[str] = Query(None), _user: str = Depends(require_auth)):
    data = await cache.get_all()
    merged = data.get("merged_programs", [])
    
    if nivel and nivel.upper() != "TODOS":
        nivel = nivel.upper()
        merged = [p for p in merged if p.get("nivel") == nivel]
        
        totals = {
            "solicitados": sum(p.get("solicitados", 0) for p in merged),
            "admitidos": sum(p.get("admitidos", 0) for p in merged),
            "pagados": sum(p.get("pagados", 0) for p in merged),
            "metas": sum(p.get("meta", 0) for p in merged),
        }
        return {"estados_by_programa": merged, "totals": totals, "trends": {}, "admitidos_status": {}, "estados_gestion": []}

    return {
        "estados_by_programa": merged,
        "totals": data.get("totals", {}),
        "trends": data.get("trends", {}),
        "admitidos_status": {},
        "estados_gestion": []
    }


@router.get("/no-util")
async def get_no_util(nivel: Optional[str] = Query(None), _user: str = Depends(require_auth)):
    from database import fetch_all
    data_cache = await cache.get_all()

    rows = []
    try:
        # Query agg_no_utiles directly so we get all subcategories
        if nivel and nivel.upper() != "TODOS":
            target_nivel = nivel.upper()
            programs_of_level = [p["programa"] for p in data_cache.get("merged_programs", []) if p.get("nivel") == target_nivel]

            if not programs_of_level:
                return {"no_util": [], "no_util_total": 0, "trends": {}}

            placeholders = ",".join(f"${i+1}" for i in range(len(programs_of_level)))
            query = f"""
                SELECT descripcion_sub,
                       SUM(leads_no_utiles) AS leads
                FROM agg_no_utiles
                WHERE UPPER(TRIM(programa)) IN ({placeholders})
                GROUP BY descripcion_sub
                ORDER BY leads DESC
            """
            rows = await fetch_all(query, *programs_of_level)
        else:
            rows = await fetch_all(
                "SELECT descripcion_sub, SUM(leads_no_utiles) AS leads FROM agg_no_utiles GROUP BY descripcion_sub ORDER BY leads DESC"
            )
    except Exception as e:
        print(f"[no-util] agg_no_utiles query failed ({e}), falling back to cache")
        rows = []

    # Fallback: if query returned nothing, use the cached dim_contactos data
    if not rows:
        cached_no_util = data_cache.get("no_util", [])
        total_fb = sum(item.get("leads", 0) for item in cached_no_util)
        result_fb = [
            {
                "subcategoria": item.get("descripcion_sub", ""),
                "leads": item.get("leads", 0),
                "leads_7d": item.get("leads_7d", 0),
                "leads_14d": item.get("leads_14d", 0),
                "porcentaje": round(item.get("leads", 0) / total_fb * 100, 2) if total_fb else 0,
            }
            for item in cached_no_util
        ]
        return {"no_util": result_fb, "no_util_total": total_fb, "trends": data_cache.get("trends", {})}

    total = sum(int(r.get("leads") or 0) for r in rows)

    result = [
        {
            "subcategoria": r.get("descripcion_sub", ""),
            "leads": int(r.get("leads") or 0),
            "leads_7d": 0,
            "leads_14d": 0,
            "porcentaje": round(int(r.get("leads") or 0) / total * 100, 2) if total else 0,
        }
        for r in rows
    ]

    return {"no_util": result, "no_util_total": total, "trends": data_cache.get("trends", {})}


@router.get("/export-no-util")
async def export_no_util(
    nivel: Optional[str] = Query(None),
    _user: str = Depends(require_auth)
):
    from database import fetch_all
    data_cache = await cache.get_all()
    
    # Logic similar to get_no_util but for full export
    rows = []
    try:
        if nivel and nivel.upper() != "TODOS":
            target_nivel = nivel.upper()
            programs_of_level = [p["programa"] for p in data_cache.get("merged_programs", []) if p.get("nivel") == target_nivel]
            if programs_of_level:
                placeholders = ",".join(f"${i+1}" for i in range(len(programs_of_level)))
                query = f"""
                    SELECT descripcion_sub as "MOTIVO",
                           SUM(leads_no_utiles) AS "CANTIDAD LEADS"
                    FROM agg_no_utiles
                    WHERE UPPER(TRIM(programa)) IN ({placeholders})
                    GROUP BY descripcion_sub
                    ORDER BY "CANTIDAD LEADS" DESC
                """
                rows = await fetch_all(query, *programs_of_level)
        else:
            rows = await fetch_all(
                'SELECT descripcion_sub as "MOTIVO", SUM(leads_no_utiles) AS "CANTIDAD LEADS" FROM agg_no_utiles GROUP BY descripcion_sub ORDER BY "CANTIDAD LEADS" DESC'
            )
    except:
        rows = []

    if not rows:
        # Fallback to cache if table query fails
        cached_no_util = data_cache.get("no_util", [])
        rows = [{"MOTIVO": item.get("descripcion_sub"), "CANTIDAD LEADS": item.get("leads")} for item in cached_no_util]

    df = pd.DataFrame(rows)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='No Utiles')
        apply_excel_style(writer.sheets['No Utiles'])

    output.seek(0)
    filename = f"Reporte_No_Utiles_{nivel if nivel else 'GLOBAL'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@router.get("/no-util-csv")
async def download_no_util_csv(_user: str = Depends(require_auth)):
    """Download the full agg_no_utiles_completo table as a CSV file."""
    from database import fetch_all
    try:
        rows = await fetch_all("SELECT * FROM agg_no_utiles_completo")
    except Exception as e:
        print(f"[no-util-csv] Error fetching agg_no_utiles_completo: {e}")
        return Response(content=f"Error al acceder a la tabla: {e}", media_type="text/plain", status_code=500)
    
    if not rows:
        return Response(content="Sin datos en agg_no_utiles_completo", media_type="text/plain")
    
    headers_list = list(rows[0].keys())
    lines = [",".join(headers_list)]
    for row in rows:
        values = []
        for h in headers_list:
            val = str(row.get(h, "") if row.get(h) is not None else "")
            if "," in val or '"' in val:
                val = f'"{val.replace(chr(34), chr(34)+chr(34))}"'
            values.append(val)
        lines.append(",".join(values))
    csv_content = "\n".join(lines)
    
    return Response(
        content=csv_content.encode("utf-8-sig"),
        media_type="text/csv",
        headers={"Content-Disposition": 'attachment; filename="agg_no_utiles_completo.csv"'},
    )


@router.get("/admitidos")
async def get_admitidos(_user: str = Depends(require_auth)):
    # Returns empty logic since fact_unab_sheet_admitidos is dropped
    return {
        "status": {},
        "by_programa": [],
        "motivos": [],
    }


@router.get("/leads")
async def get_leads(
    page: int = Query(1, ge=1),
    per_page: int = Query(25, ge=1, le=100),
    search: Optional[str] = Query(None),
    base: Optional[str] = Query(None),
    programa: Optional[str] = Query(None),
    nivel: Optional[str] = Query(None),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[str] = Query(None),
    fecha_fin: Optional[str] = Query(None),
    _user: str = Depends(require_auth),
):
    from database import fetch_all, fetch_one
    from mapping import mapping
    
    where_clauses = ["1=1"]
    args = []
    
    if search:
        where_clauses.append(f"(txtnombreapellido ILIKE ${len(args)+1} OR emlmail ILIKE ${len(args)+1} OR teltelefono ILIKE ${len(args)+1})")
        args.append(f"%{search}%")
        
    if base:
        where_clauses.append(f"base = ${len(args)+1}")
        args.append(base)
        
    if programa:
        where_clauses.append(f"txtprogramainteres ILIKE ${len(args)+1}")
        args.append(f"%{programa}%")
        
    if estado:
        where_clauses.append(f"ultima_mejor_subcat_string = ${len(args)+1}")
        args.append(estado)
        
    if fecha_inicio:
        where_clauses.append(f"fecha_a_utilizar >= ${len(args)+1}")
        args.append(fecha_inicio)
        
    if fecha_fin:
        where_clauses.append(f"fecha_a_utilizar <= ${len(args)+1}")
        args.append(fecha_fin)
        
    if nivel and nivel.upper() != "TODOS":
        target_nivel = nivel.upper()
        # Use the programs already classified in the cache to ensure consistency
        data = await cache.get_all()
        programs_of_level = [p["programa"] for p in data.get("merged_programs", []) if p.get("nivel") == target_nivel]
        
        if programs_of_level:
            placeholders = ",".join(f"${len(args)+i+1}" for i in range(len(programs_of_level)))
            where_clauses.append(f"UPPER(TRIM(txtprogramainteres)) IN ({placeholders})")
            args.extend(programs_of_level)
        else:
            # If no programs match this level, force no results
            where_clauses.append("1=0")
        
    where_sql = " AND ".join(where_clauses)
    
    count_query = f"SELECT COUNT(*) as total FROM dim_contactos WHERE {where_sql}"
    total_row = await fetch_one(count_query, *args)
    total = total_row["total"] if total_row else 0
    
    offset = (page - 1) * per_page
    data_query = f"""
        SELECT 
            idinterno, txtnombreapellido, emlmail, teltelefono, 
            feccreacionoportunidad, txtprogramainteres, base,
            descrip_subcat, ultima_mejor_subcat_string, cant_toques_call_crm, fecha_a_utilizar
        FROM dim_contactos
        WHERE {where_sql}
        ORDER BY fecha_a_utilizar DESC NULLS LAST
        LIMIT {per_page} OFFSET {offset}
    """
    
    rows = await fetch_all(data_query, *args)
    
    return {
        "data": rows,
        "total": total,
        "page": page,
        "per_page": per_page,
    }


@router.get("/bases")
async def get_bases(_user: str = Depends(require_auth)):
    from database import fetch_all
    query = "SELECT DISTINCT base FROM dim_contactos WHERE base IS NOT NULL ORDER BY base"
    rows = await fetch_all(query)
    return [r["base"] for r in rows]

@router.get("/estados-gestion")
async def get_estados_gestion(_user: str = Depends(require_auth)):
    from database import fetch_all
    query = "SELECT DISTINCT ultima_mejor_subcat_string FROM dim_contactos WHERE ultima_mejor_subcat_string IS NOT NULL ORDER BY ultima_mejor_subcat_string"
    rows = await fetch_all(query)
    return [r["ultima_mejor_subcat_string"] for r in rows]

@router.get("/meta")
async def get_meta(_user: str = Depends(require_auth)):
    data = await cache.get_all()
    return {
        "fecha_actualizacion": data.get("fecha_actualizacion", ""),
        "last_refresh": cache.last_refresh.isoformat() if cache.last_refresh else None,
        "total_leads": data.get("total_leads", 0),
    }

@router.post("/refresh")
async def manual_refresh(_user: str = Depends(require_auth)):
    await cache.refresh()
    return {"status": "success", "last_refresh": cache.last_refresh.isoformat()}
